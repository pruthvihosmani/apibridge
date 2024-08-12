import socket
import json
import openpyxl
import time

# Configuration
HOST = '127.0.0.1'  # Localhost
PORT = 30001        # The port APIBridge is listening on
EXCEL_FILE = 'live_positions.xlsx'

# Function to connect to APIBridge
def connect_to_apibridge():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect((HOST, PORT))
        print("Connected to API Bridge")
        return s
    except Exception as e:
        print(f"Failed to connect to API Bridge: {e}")
        return None

# Function to request live positions from APIBridge
def get_live_positions(socket_conn):
    try:
        request_data = json.dumps({
            "action": "GET_POSITIONS"
        })
        socket_conn.sendall(request_data.encode('utf-8'))

        # Receive the response
        response_data = socket_conn.recv(4096)  # Increase buffer size if needed
        positions = json.loads(response_data.decode('utf-8'))
        return positions
    except Exception as e:
        print(f"Failed to retrieve live positions: {e}")
        return None

# Function to write positions to Excel
def write_positions_to_excel(positions):
    # Load or create Excel workbook
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select or create a worksheet
    sheet = workbook.active
    sheet.title = "Live Positions"

    # Clear existing data in the worksheet
    sheet.delete_rows(1, sheet.max_row)

    # Write header
    headers = ["Symbol", "Qty", "Buy Price", "LTP", "P&L", "SL Trigger", "SL Price"]
    sheet.append(headers)

    # Write position data
    for position in positions.get('positions', []):
        row = [
            position.get('symbol', ''),
            position.get('quantity', ''),
            position.get('buy_price', ''),
            position.get('ltp', ''),
            position.get('pnl', ''),
            position.get('sl_trigger', ''),
            position.get('sl_price', '')
        ]
        sheet.append(row)

    # Save the Excel file
    workbook.save(EXCEL_FILE)
    print(f"Live positions updated in {EXCEL_FILE}")

# Main function to run the script
def main():
    # Connect to APIBridge
    socket_conn = connect_to_apibridge()
    if socket_conn is None:
        print("Exiting...")
        return

    try:
        while True:
            # Get live positions
            positions = get_live_positions(socket_conn)
            if positions:
                write_positions_to_excel(positions)

            # Wait for a few seconds before fetching again (e.g., every 10 seconds)
            time.sleep(10)
    finally:
        socket_conn.close()

if __name__ == "__main__":
    main()
